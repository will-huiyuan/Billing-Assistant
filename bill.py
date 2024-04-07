
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import datetime
import os

def get_file_path():
    # 检测操作系统
    if os.name == 'nt':  # Windows系统
        return r'如果你是windows系统，请将记账表格的路径复制到这里'
    elif os.name == 'posix':  # macOS或Linux系统
        return r'如果你是macOS或Linux系统，请将记账表格的路径复制到这里'
    else:
        raise Exception('Unsupported Operating System')

replace_merchant = {
    # 这里是商家名称的替换字典，你可以强制将某个商家的名称替换为另一个名称
    "bines_service":"MajesThé",
    "商家名称":"新商家名称"
}
merchant_type_match = {
    # 这里是商家名称和类型的匹配字典，你可以强制将某个商家的类型设置为另一个类型
    "MajesThé": "食品",
    "商家名称": "新商家类型"
}

def append_data_to_excel(data, file_name):
    # open the workbook
    try:
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
    except (FileNotFoundError, InvalidFileException):
        print("Error")
        exit(1)
        # workbook = Workbook()
        # sheet = workbook.active
    # manual replacement
    data["merchant"] = data["merchant"].lower()
    if (data["merchant"]) in replace_merchant:
        data["merchant"] = replace_merchant[data["merchant"].lower()]
    # if date is empty, use the date today
    if data["date"] is None:
        data["date"] = datetime.date.today().strftime("%Y-%m-%d")
    # if type is empty, use "食品"
    if data["type"] is None:
        data["type"] = '食品'
    # if merchant match dict, use dict type
    if data["merchant"] in merchant_type_match:
        data["type"] = merchant_type_match[data["merchant"]]
    # append the data
    sheet.append([data["date"], data["price"], data["merchant"], data["card"],data["type"]])
    workbook.save(filename=file_name)
    workbook.close()
    # message
    print(data_dict,"\nsuccess")

#更改这个字典来添加你的数据
data_dict = \
{
"merchant": "name",
# 日期为空时，将使用今天的日期
"date": None,
"price": 17.25,
"card": "1234",
"type": "娱乐"
}











append_data_to_excel(data_dict, get_file_path())


